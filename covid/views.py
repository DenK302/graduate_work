# import os, zipfile
#
# from django.contrib.auth.decorators import permission_required
import os
import glob

from django.contrib.auth.decorators import login_required, permission_required
# from django.db.models import FloatField
# from django.db.models import Q, Sum, Count
# from django.db.models.functions import TruncYear, Cast
from django.http import Http404, HttpResponseRedirect, HttpResponseNotFound, HttpResponse, FileResponse
from django.shortcuts import render
from django.urls import reverse
from django.views import generic
# from django.views.generic import ListView
from openpyxl import load_workbook

from covid19db import settings
from .models import Employer
import datetime
from datetime import date, timedelta
from docxtpl import DocxTemplate
import urllib.parse


#
#

@login_required
def EmployerListView(request):
    employer_list = []
    group_name = str()
    department_list = Employer.employers_department
    for group in request.user.groups.all():
        group_name = str(group.name)
        for index, name in Employer.employers_department:
            if name == group.name:
                employer_list += (Employer.objects.filter(department=index))

    return render(
        request,
        'index.html',
        context={'employer_list': employer_list, 'group_name': group_name, 'department_list': department_list},
    )


@login_required
def EmployerFilter(request, index_group):
    group_name = str()
    department_list = Employer.employers_department
    employer_list = Employer.objects.filter(department=index_group)
    for index, name in Employer.employers_department:
        if index_group == index:
            group_name = name
    return render(
        request,
        'index.html',
        context={'employer_list': employer_list, 'group_name_filter': group_name, 'department_list': department_list},
    )


@login_required
def EditEmployer(request, pk):
    try:
        employer = Employer.objects.get(pk=pk)
        if request.POST:
            if request.POST.get("fullname"):
                employer.fullname = request.POST.get("fullname")
            if request.POST.get("birthdate"):
                employer.birthdate = request.POST.get("birthdate")
            if request.POST.get("liveplace"):
                employer.liveplace = request.POST.get("liveplace")
            if request.POST.get("department"):
                employer.department = request.POST.get("department")
            if request.POST.get("position"):
                employer.position = request.POST.get("position")
            if request.POST.get("datefirst") == '':
                employer.datefirst = None
            else:
                employer.datefirst = request.POST.get("datefirst")
            if request.POST.get("datesecond") == '':
                employer.datesecond = None
            else:
                employer.datesecond = request.POST.get("datesecond")
            if request.POST.get("vactinationfull") == 'on':
                employer.vactinationfull = True
            else:
                employer.vactinationfull = False
            if request.POST.get("sertification") == "on":
                employer.sertification = True
                employer.sertificateid = request.POST.get("sertificateid")
            else:
                employer.sertification = False
                employer.sertificateid = None
            employer.vactinationplace1 = request.POST.get("vactinationplace1")
            employer.vactinationplace2 = request.POST.get("vactinationplace2")
            if request.POST.get("wantmake") == "on":
                employer.wantmake = True
                employer.wantmaketype = request.POST.get("wantmaketype")
            else:
                employer.wantmake = False
                employer.wantmaketype = ""
            employer.contraindications = request.POST.get("contraindications")
            employer.save()
            return HttpResponseRedirect(reverse('employers'))
        else:
            return render(request, "covid/employer_detail.html",
                          {"employer": employer, "department_list": Employer.employers_department})
    except Employer.DoesNotExist:
        return HttpResponseNotFound("<h2>Сотрудник не найден</h2>")


@permission_required('covid.delete_employer')
def DeleteEmployer(request, pk):
    try:
        employer = Employer.objects.get(pk=pk)
        employer.delete()
        return HttpResponseRedirect(reverse('employers'))
    except Employer.DoesNotExist:
        return HttpResponseNotFound("<h2>Сотрудник не найден</h2>")


@permission_required('covid.delete_employer')
def CreateEmployer(request):
    if request.POST:
        if request.POST['datefirst'] == '':
            datefirst = None
        else:
            datefirst = request.POST['datefirst']
        if request.POST['datesecond'] == '':
            datesecond = None
        else:
            datesecond = request.POST['datesecond']
        if 'vactinationfull' in request.POST:
            vactinationfull = True
        else:
            vactinationfull = False
        if 'sertification' in request.POST:
            sertification = True
            sertificateid = request.POST['sertificateid']
        else:
            sertification = False
            sertificateid = None
        if 'wantmake' in request.POST:
            wantmake = True
            wantmaketype = request.POST['wantmaketype']
        else:
            wantmake = False
            wantmaketype = ""
        employer = Employer(
            fullname=request.POST['fullname'],
            birthdate=request.POST['birthdate'],
            liveplace=request.POST['liveplace'],
            department=request.POST['department'],
            position=request.POST['position'],
            datefirst=datefirst,
            datesecond=datesecond,
            vactinationfull=vactinationfull,
            sertification=sertification,
            sertificateid=sertificateid,
            vactinationplace1=request.POST['vactinationplace1'],
            vactinationplace2=request.POST['vactinationplace2'],
            wantmake=wantmake,
            wantmaketype=wantmaketype,
            contraindications=request.POST['contraindications']
        )
        employer.save()
        return HttpResponseRedirect(reverse('employers'))
    else:
        return render(request, "covid/employer_create.html", {"department_list": Employer.employers_department})


@permission_required('covid.delete_employer')
def ParseEmployer(request):
    if request.POST:
        employer_list = []
        # employer_list = Employer.objects.values_list("fullname", flat=True)
        for employer in Employer.objects.values_list("fullname", flat=True):
            employer_list.append(employer)
        wb = load_workbook(request.FILES['file'])
        steck = ['Административно-хозяйственная часть',
                 'Аспирантура',
                 'Библиотека',
                 'деканат факультета мониторинга окружающей среды',
                 'деканат факультета повышения квалификации и переподготовки',
                 'деканат факультета экологической медицины',
                 'Директорат',
                 'кафедра дополнительного образования',
                 'кафедра иммунологии',
                 'кафедра информационных технологий в экологии и медицине',
                 'кафедра лингвистических дисциплин и межкультурных коммуникаций',
                 'кафедра общей биологии и генетики',
                 'кафедра общей и медицинской физики',
                 'кафедра социально-гуманитарных наук и устойчивого развития',
                 'кафедра физического воспитания',
                 'кафедра экологического мониторинга и менеджмента',
                 'кафедра экологической медицины и радиобиологии',
                 'кафедра экологической химии и биохимии',
                 'кафедра энергоэффективных технологий',
                 'кафедра ядерной и радиационной безопасности',
                 'Научно-исследовательский сектор',
                 'Отдел бухгалтерского учета, отчетности и финансирования',
                 'Отдел воспитательной работы с молодежью',
                 'Отдел закупок и материально-технического снабжения',
                 'Отдел кадровой и организационной работы',
                 'Отдел международных связей',
                 'Отдел технических средств обучения и коммуникаций',
                 'Планово-экономический сектор',
                 'Спортивный клуб',
                 'учебная лаборатория экологической биотехнологии',
                 'Учебно-методическая лаборатория инновационных технологий образования',
                 'учебно-методическая лаборатория экологического образования',
                 'Учебно-методический отдел',
                 'Учебный корпус (ул.Ботаническая, 15)',
                 'Учебный корпус № 1 (Долгобродская, 23)',
                 'Факультет экологической медицины']
        sheet = wb["TDSheet"]
        i = 1
        j = 1
        currentDepartment = ""
        while sheet.cell(row=i, column=1).value != "Итого":
            i += 1
            if sheet.cell(row=i, column=1).value == "Итого":
                break
            if sheet.cell(row=i, column=1).value in steck:
                currentDepartment = sheet.cell(row=i, column=1).value
                continue
            if (sheet.cell(row=i, column=1).value == None) | (sheet.cell(row=i, column=1).value == "Да") | (
                    sheet.cell(row=i, column=1).value == "III группа"):
                continue
            if currentDepartment != "":
                for index, name in Employer.employers_department:
                    if name == currentDepartment:
                        emplpoyer = Employer(
                            fullname=sheet.cell(row=i, column=1).value,
                            birthdate=sheet.cell(row=i, column=5).value,
                            liveplace=sheet.cell(row=i, column=6).value,
                            department=index,
                            position=sheet.cell(row=i, column=2).value,
                            datefirst=None,
                            datesecond=None,
                            vactinationfull=False,
                            sertification=False,
                            sertificateid="",
                            vactinationplace1="",
                            vactinationplace2="",
                            wantmake=False,
                            wantmaketype="",
                            contraindications=""
                        )
                        if emplpoyer.fullname not in employer_list:
                            emplpoyer.save()
        return HttpResponseRedirect(reverse('employers'))
    else:
        department_list = Employer.employers_department
        return render(request, "covid/employer_parse.html", {'department_list': department_list})


@permission_required('covid.delete_employer')
def CreateReport(request):
    department_list = Employer.employers_department
    return render(request, "covid/employer_report.html", {'department_list': department_list})


@permission_required('covid.delete_employer')
def TuesdayThursday(request):
    countemployer = Employer.objects.count()
    currentdate = datetime.datetime.strptime(str(date.today()), "%Y-%m-%d").strftime("%d.%m.%Y")
    countemployervactinated = Employer.objects.filter(vactinationfull=True).count()
    vactinationpercent = round((countemployervactinated / countemployer) * 100, 1)
    doc = DocxTemplate("docs/Otchet o vakcinacii (vtornik, chetverg) shablon.docx")
    context = {'countEmployer': str(countemployer), 'VactinatedByDate': str(countemployervactinated),
               "VactinatedPercent": str(vactinationpercent), "Date": str(currentdate)}
    doc.render(context)
    doc.save("docs/Otchet o vakcinacii (vtornik, chetverg) na " + str(currentdate) + ".docx")
    response = FileResponse(
        open("docs/Otchet o vakcinacii (vtornik, chetverg) na " + str(currentdate) + ".docx", 'rb'))
    return response


@permission_required('covid.delete_employer')
def Monday(request):
    countemployer = Employer.objects.count()
    currentdate = datetime.datetime.strptime(str(date.today()), "%Y-%m-%d").strftime("%d.%m.%Y")
    vactinationCountFirst = Employer.objects.filter(datefirst__gte=date.today() - timedelta(days=182)).count()
    vactinationCountSecond = Employer.objects.filter(datesecond__gte=date.today() - timedelta(days=182)).count()
    percentVactinationFirst = round((vactinationCountFirst / countemployer) * 100, 1)
    percentVactinationSecond = round((vactinationCountSecond / countemployer) * 100, 1)
    countContraindications = Employer.objects.exclude(contraindications="").count()
    countWanted = Employer.objects.filter(wantmake=True).count()
    countKravira = Employer.objects.filter(vactinationplace2="УЗ «Кравира»").count()
    count14 = Employer.objects.filter(vactinationplace2="14-я городская поликлиника").count()
    count9 = Employer.objects.filter(vactinationplace2="9-я городская поликлиника").count()
    countMTZ = Employer.objects.filter(vactinationplace2="Медцентр ОАО «МТЗ»").count()
    doc = DocxTemplate("docs/Otchet o vakcinacii (ponedelnik) shablon.docx")
    context = {'countEmployer': str(countemployer), 'vactinationCountFirst': str(vactinationCountFirst),
               'percentVactinationFirst': str(percentVactinationFirst),
               'percentVactinationSecond': str(percentVactinationSecond),
               'countContraindications': str(countContraindications), 'countWanted': str(countWanted),
               'countKravira': countKravira, 'count14': count14, 'count9': count9, 'countMTZ': countMTZ,
               "vactinationCountSecond": str(vactinationCountSecond), "Date": str(currentdate)}
    doc.render(context)
    doc.save("docs/Otchet o vakcinacii (ponedelnik) na " + str(currentdate) + ".docx")
    response = FileResponse(
        open("docs/Otchet o vakcinacii (ponedelnik) na " + str(currentdate) + ".docx", 'rb'))
    return response


@permission_required('covid.delete_employer')
def Monday16(request):
    countemployer = Employer.objects.count()
    currentdate = datetime.datetime.strptime(str(date.today()), "%Y-%m-%d").strftime("%d.%m.%Y")
    employerVactinated = Employer.objects.filter(vactinationfull=True).count()
    percentVactinated = round(employerVactinated / countemployer * 100, 1)
    doc = DocxTemplate("docs/Otchet o vakcinacii (ponedelnik do 16.00) shablon.docx")
    context = {'countEmployer': str(countemployer), "Date": str(currentdate), 'employerVactinated': employerVactinated,
               'percentVactinated': percentVactinated}
    doc.render(context)
    doc.save("docs/Otchet o vakcinacii (ponedelnik do 16.00) na " + str(currentdate) + ".docx")
    response = FileResponse(
        open("docs/Otchet o vakcinacii (ponedelnik do 16.00) na " + str(currentdate) + ".docx", 'rb'))
    return response


@permission_required('covid.delete_employer')
def EveryDay(request):
    countemployer = Employer.objects.count()
    currentdate = datetime.datetime.strptime(str(date.today()), "%Y-%m-%d").strftime("%d.%m.%Y")
    countemployervactinated = Employer.objects.filter(vactinationfull=True).count()
    countWanted = Employer.objects.filter(wantmake=True).count()
    countKravira = Employer.objects.filter(vactinationplace2="УЗ «Кравира»").count()
    count14 = Employer.objects.filter(vactinationplace2="14-я городская поликлиника").count()
    count9 = Employer.objects.filter(vactinationplace2="9-я городская поликлиника").count()
    countMTZ = Employer.objects.filter(vactinationplace2="Медцентр ОАО «МТЗ»").count()
    countpartiz = countKravira + count14 + count9 + countMTZ
    countanother = countemployervactinated - countpartiz
    doc = DocxTemplate("docs/Ezhednevnyj otchet o vakcinacii shablon.docx")
    context = {'countEmployer': str(countemployer), 'countpartiz': str(countpartiz),
               'countanother': str(countanother), 'countWanted': str(countWanted), "Date": str(currentdate)}
    doc.render(context)
    doc.save("docs/Ezhednevnyj otchet o vakcinacii na " + str(currentdate) + ".docx")
    response = FileResponse(
        open("docs/Ezhednevnyj otchet o vakcinacii na " + str(currentdate) + ".docx", 'rb'))
    return response
